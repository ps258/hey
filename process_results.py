#!/usr/bin/env python3
"""
Process CSV output from hey load testing tool and generate Excel spreadsheet with charts.

Expected CSV format:
timestamp,avg_response_time,p95_response_time,requests_per_second

Usage:
    python process_results.py input.csv output.xlsx
"""

import sys
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def read_csv_data(csv_file):
    """Read CSV file and return data as list of dictionaries."""
    data = []
    try:
        with open(csv_file, 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                data.append({
                    'timestamp': row['timestamp'],
                    'avg_response_time': float(row['avg_response_time']),
                    'p95_response_time': float(row['p95_response_time']),
                    'requests_per_second': float(row['requests_per_second'])
                })
        return data
    except FileNotFoundError:
        print(f"Error: File '{csv_file}' not found.")
        sys.exit(1)
    except KeyError as e:
        print(f"Error: Missing expected column {e} in CSV file.")
        print("Expected columns: timestamp,avg_response_time,p95_response_time,requests_per_second")
        sys.exit(1)
    except ValueError as e:
        print(f"Error: Invalid numeric value in CSV file: {e}")
        sys.exit(1)


def calculate_statistics(data):
    """Calculate overall statistics from the data."""
    if not data:
        return {}
    
    avg_times = [row['avg_response_time'] for row in data]
    p95_times = [row['p95_response_time'] for row in data]
    rps_values = [row['requests_per_second'] for row in data]
    
    stats = {
        'total_samples': len(data),
        'avg_response_time_mean': sum(avg_times) / len(avg_times),
        'avg_response_time_min': min(avg_times),
        'avg_response_time_max': max(avg_times),
        'p95_response_time_mean': sum(p95_times) / len(p95_times),
        'p95_response_time_min': min(p95_times),
        'p95_response_time_max': max(p95_times),
        'rps_mean': sum(rps_values) / len(rps_values),
        'rps_min': min(rps_values),
        'rps_max': max(rps_values),
    }
    
    return stats


def create_excel_with_charts(data, stats, output_file):
    """Create Excel workbook with data and charts."""
    wb = Workbook()
    
    # Create Data sheet
    ws_data = wb.active
    ws_data.title = "Data"
    
    # Write headers with styling
    headers = ['Timestamp', 'Avg Response Time (s)', 'P95 Response Time (s)', 'Requests/Second']
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, start=1):
        cell = ws_data.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row_idx, row_data in enumerate(data, start=2):
        ws_data.cell(row=row_idx, column=1, value=row_data['timestamp'])
        ws_data.cell(row=row_idx, column=2, value=row_data['avg_response_time'])
        ws_data.cell(row=row_idx, column=3, value=row_data['p95_response_time'])
        ws_data.cell(row=row_idx, column=4, value=row_data['requests_per_second'])
    
    # Auto-adjust column widths
    for col in range(1, 5):
        ws_data.column_dimensions[get_column_letter(col)].width = 20
    
    # Create Statistics sheet
    ws_stats = wb.create_sheet("Statistics")
    ws_stats['A1'] = "Metric"
    ws_stats['B1'] = "Value"
    ws_stats['A1'].font = header_font
    ws_stats['B1'].font = header_font
    ws_stats['A1'].fill = header_fill
    ws_stats['B1'].fill = header_fill
    
    stats_rows = [
        ("Total Samples", stats['total_samples']),
        ("", ""),
        ("Average Response Time (s)", ""),
        ("  Mean", stats['avg_response_time_mean']),
        ("  Min", stats['avg_response_time_min']),
        ("  Max", stats['avg_response_time_max']),
        ("", ""),
        ("P95 Response Time (s)", ""),
        ("  Mean", stats['p95_response_time_mean']),
        ("  Min", stats['p95_response_time_min']),
        ("  Max", stats['p95_response_time_max']),
        ("", ""),
        ("Requests per Second", ""),
        ("  Mean", stats['rps_mean']),
        ("  Min", stats['rps_min']),
        ("  Max", stats['rps_max']),
    ]
    
    for row_idx, (metric, value) in enumerate(stats_rows, start=2):
        ws_stats.cell(row=row_idx, column=1, value=metric)
        if isinstance(value, (int, float)):
            ws_stats.cell(row=row_idx, column=2, value=round(value, 4))
        else:
            ws_stats.cell(row=row_idx, column=2, value=value)
    
    ws_stats.column_dimensions['A'].width = 30
    ws_stats.column_dimensions['B'].width = 15
    
    # Create Charts sheet
    ws_charts = wb.create_sheet("Charts")
    
    # Chart 1: Response Times
    chart1 = LineChart()
    chart1.title = "Response Times Over Time"
    chart1.style = 10
    chart1.y_axis.title = "Response Time (seconds)"
    chart1.x_axis.title = "Sample"
    chart1.height = 10
    chart1.width = 20
    
    # Add data series for response times
    data_ref = Reference(ws_data, min_col=2, min_row=1, max_row=len(data) + 1)
    chart1.add_data(data_ref, titles_from_data=True)
    
    data_ref = Reference(ws_data, min_col=3, min_row=1, max_row=len(data) + 1)
    chart1.add_data(data_ref, titles_from_data=True)
    
    # Set categories (x-axis)
    cats = Reference(ws_data, min_col=1, min_row=2, max_row=len(data) + 1)
    chart1.set_categories(cats)
    
    ws_charts.add_chart(chart1, "A1")
    
    # Chart 2: Requests per Second
    chart2 = LineChart()
    chart2.title = "Requests per Second Over Time"
    chart2.style = 12
    chart2.y_axis.title = "Requests/Second"
    chart2.x_axis.title = "Sample"
    chart2.height = 10
    chart2.width = 20
    
    # Add data series for RPS
    data_ref = Reference(ws_data, min_col=4, min_row=1, max_row=len(data) + 1)
    chart2.add_data(data_ref, titles_from_data=True)
    chart2.set_categories(cats)
    
    ws_charts.add_chart(chart2, "A25")
    
    # Save workbook
    try:
        wb.save(output_file)
        print(f"Successfully created Excel file: {output_file}")
        print(f"  - Data sheet with {len(data)} samples")
        print(f"  - Statistics sheet with summary metrics")
        print(f"  - Charts sheet with visualizations")
    except PermissionError:
        print(f"Error: Cannot write to '{output_file}'. File may be open in another program.")
        sys.exit(1)


def main():
    """Main function to process CSV and generate Excel."""
    if len(sys.argv) != 3:
        print("Usage: python process_results.py input.csv output.xlsx")
        print("\nExpected CSV format:")
        print("timestamp,avg_response_time,p95_response_time,requests_per_second")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2]
    
    if not output_file.endswith('.xlsx'):
        output_file += '.xlsx'
    
    print(f"Reading data from: {input_file}")
    data = read_csv_data(input_file)
    
    if not data:
        print("Error: No data found in CSV file.")
        sys.exit(1)
    
    print(f"Processing {len(data)} samples...")
    stats = calculate_statistics(data)
    
    print(f"Generating Excel file: {output_file}")
    create_excel_with_charts(data, stats, output_file)
    
    print("\nSummary Statistics:")
    print(f"  Average Response Time: {stats['avg_response_time_mean']:.4f}s")
    print(f"  P95 Response Time: {stats['p95_response_time_mean']:.4f}s")
    print(f"  Average RPS: {stats['rps_mean']:.2f}")


if __name__ == "__main__":
    main()